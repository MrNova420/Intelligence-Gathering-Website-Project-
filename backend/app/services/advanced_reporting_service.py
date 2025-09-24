"""
Advanced Reporting and Analytics Service
=======================================

Enterprise-grade reporting capabilities with:
- Multi-format report generation (PDF, Excel, HTML, JSON)
- Advanced data visualization and charts  
- Executive dashboards and summaries
- Automated report scheduling
- Custom report templates
- Data export and integration APIs
"""

import asyncio
import logging
import json
import base64
from datetime import datetime, timedelta
from typing import Dict, Any, List, Optional, Union
from dataclasses import dataclass, field
from enum import Enum
import statistics
import io

logger = logging.getLogger(__name__)

# Mock imports for optional dependencies
try:
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    from matplotlib.backends.backend_agg import FigureCanvasAgg
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False
    plt = None

try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ReportFormat(str, Enum):
    """Supported report formats"""
    PDF = "pdf"
    HTML = "html"
    JSON = "json"
    CSV = "csv"
    EXCEL = "excel"
    XML = "xml"


class ReportType(str, Enum):
    """Types of reports"""
    EXECUTIVE_SUMMARY = "executive_summary"
    DETAILED_ANALYSIS = "detailed_analysis"
    TECHNICAL_REPORT = "technical_report"
    DASHBOARD = "dashboard"
    CUSTOM = "custom"


class VisualizationType(str, Enum):
    """Types of data visualizations"""
    LINE_CHART = "line_chart"
    BAR_CHART = "bar_chart"
    PIE_CHART = "pie_chart"
    SCATTER_PLOT = "scatter_plot"
    HEATMAP = "heatmap"
    TABLE = "table"


@dataclass
class ReportSection:
    """Individual report section"""
    title: str
    content: str
    section_type: str = "text"
    data: Optional[Dict[str, Any]] = None
    visualization: Optional[Dict[str, Any]] = None
    order: int = 0


@dataclass
class ReportTemplate:
    """Report template definition"""
    template_id: str
    name: str
    description: str
    report_type: ReportType
    sections: List[str] = field(default_factory=list)
    parameters: Dict[str, Any] = field(default_factory=dict)
    styling: Dict[str, Any] = field(default_factory=dict)


@dataclass
class ReportRequest:
    """Report generation request"""
    report_id: str
    template_id: str
    report_type: ReportType
    format: ReportFormat
    title: str
    parameters: Dict[str, Any] = field(default_factory=dict)
    data_sources: List[str] = field(default_factory=list)
    include_visualizations: bool = True
    include_raw_data: bool = False
    custom_styling: Optional[Dict[str, Any]] = None


class DataVisualizationEngine:
    """Engine for creating data visualizations"""
    
    def __init__(self):
        self.logger = logging.getLogger(f"{__name__}.{self.__class__.__name__}")
    
    async def create_visualization(
        self, 
        data: Dict[str, Any], 
        viz_type: VisualizationType,
        title: str = "",
        config: Dict[str, Any] = None
    ) -> Optional[str]:
        """Create a data visualization and return as base64 encoded image"""
        
        if not MATPLOTLIB_AVAILABLE:
            self.logger.warning("Matplotlib not available, skipping visualization")
            return None
        
        try:
            config = config or {}
            
            if viz_type == VisualizationType.LINE_CHART:
                return await self._create_line_chart(data, title, config)
            elif viz_type == VisualizationType.BAR_CHART:
                return await self._create_bar_chart(data, title, config)
            elif viz_type == VisualizationType.PIE_CHART:
                return await self._create_pie_chart(data, title, config)
            elif viz_type == VisualizationType.SCATTER_PLOT:
                return await self._create_scatter_plot(data, title, config)
            else:
                self.logger.warning(f"Unsupported visualization type: {viz_type}")
                return None
                
        except Exception as e:
            self.logger.error(f"Error creating visualization: {e}")
            return None
    
    async def _create_line_chart(self, data: Dict[str, Any], title: str, config: Dict[str, Any]) -> str:
        """Create a line chart"""
        fig, ax = plt.subplots(figsize=(10, 6))
        
        # Extract data
        x_data = data.get("x", [])
        y_data = data.get("y", [])
        
        if not x_data or not y_data:
            # Generate sample data if none provided
            x_data = list(range(len(y_data))) if y_data else list(range(10))
            y_data = y_data if y_data else [i**2 for i in x_data]
        
        ax.plot(x_data, y_data, marker='o', linewidth=2)
        ax.set_title(title or "Line Chart", fontsize=14, fontweight='bold')
        ax.set_xlabel(config.get("x_label", "X Axis"))
        ax.set_ylabel(config.get("y_label", "Y Axis"))
        ax.grid(True, alpha=0.3)
        
        # Save to base64
        buffer = io.BytesIO()
        plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight')
        plt.close(fig)
        buffer.seek(0)
        
        image_base64 = base64.b64encode(buffer.read()).decode()
        return image_base64
    
    async def _create_bar_chart(self, data: Dict[str, Any], title: str, config: Dict[str, Any]) -> str:
        """Create a bar chart"""
        fig, ax = plt.subplots(figsize=(10, 6))
        
        # Extract data
        categories = data.get("categories", [])
        values = data.get("values", [])
        
        if not categories or not values:
            # Generate sample data
            categories = ["A", "B", "C", "D", "E"]
            values = [20, 35, 30, 35, 27]
        
        bars = ax.bar(categories, values, color=config.get("color", "steelblue"))
        ax.set_title(title or "Bar Chart", fontsize=14, fontweight='bold')
        ax.set_xlabel(config.get("x_label", "Categories"))
        ax.set_ylabel(config.get("y_label", "Values"))
        
        # Add value labels on bars
        for bar in bars:
            height = bar.get_height()
            ax.annotate(f'{height}',
                       xy=(bar.get_x() + bar.get_width() / 2, height),
                       xytext=(0, 3),
                       textcoords="offset points",
                       ha='center', va='bottom')
        
        # Save to base64
        buffer = io.BytesIO()
        plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight')
        plt.close(fig)
        buffer.seek(0)
        
        image_base64 = base64.b64encode(buffer.read()).decode()
        return image_base64
    
    async def _create_pie_chart(self, data: Dict[str, Any], title: str, config: Dict[str, Any]) -> str:
        """Create a pie chart"""
        fig, ax = plt.subplots(figsize=(8, 8))
        
        # Extract data
        labels = data.get("labels", [])
        sizes = data.get("sizes", [])
        
        if not labels or not sizes:
            # Generate sample data
            labels = ["Category A", "Category B", "Category C", "Category D"]
            sizes = [30, 25, 25, 20]
        
        colors = config.get("colors", plt.cm.Set3.colors)
        explode = config.get("explode", [0.05] * len(labels))
        
        wedges, texts, autotexts = ax.pie(
            sizes, labels=labels, autopct='%1.1f%%',
            colors=colors, explode=explode, shadow=True, startangle=90
        )
        
        ax.set_title(title or "Pie Chart", fontsize=14, fontweight='bold')
        
        # Save to base64
        buffer = io.BytesIO()
        plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight')
        plt.close(fig)
        buffer.seek(0)
        
        image_base64 = base64.b64encode(buffer.read()).decode()
        return image_base64
    
    async def _create_scatter_plot(self, data: Dict[str, Any], title: str, config: Dict[str, Any]) -> str:
        """Create a scatter plot"""
        fig, ax = plt.subplots(figsize=(10, 6))
        
        # Extract data
        x_data = data.get("x", [])
        y_data = data.get("y", [])
        
        if not x_data or not y_data:
            # Generate sample data
            import random
            x_data = [random.random() * 100 for _ in range(50)]
            y_data = [random.random() * 100 for _ in range(50)]
        
        ax.scatter(x_data, y_data, alpha=0.6, s=config.get("size", 50))
        ax.set_title(title or "Scatter Plot", fontsize=14, fontweight='bold')
        ax.set_xlabel(config.get("x_label", "X Axis"))
        ax.set_ylabel(config.get("y_label", "Y Axis"))
        ax.grid(True, alpha=0.3)
        
        # Save to base64
        buffer = io.BytesIO()
        plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight')
        plt.close(fig)
        buffer.seek(0)
        
        image_base64 = base64.b64encode(buffer.read()).decode()
        return image_base64


class PDFReportGenerator:
    """PDF report generation using ReportLab"""
    
    def __init__(self):
        self.logger = logging.getLogger(f"{__name__}.{self.__class__.__name__}")
    
    async def generate_pdf_report(
        self, 
        sections: List[ReportSection], 
        title: str,
        metadata: Dict[str, Any] = None
    ) -> bytes:
        """Generate a PDF report from sections"""
        
        if not REPORTLAB_AVAILABLE:
            raise ImportError("ReportLab is required for PDF generation")
        
        buffer = io.BytesIO()
        
        try:
            # Create PDF document
            doc = SimpleDocTemplate(
                buffer,
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18
            )
            
            # Define styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            
            story = []
            
            # Add title
            story.append(Paragraph(title, title_style))
            story.append(Spacer(1, 12))
            
            # Add metadata
            if metadata:
                story.append(Paragraph("Report Information", styles['Heading2']))
                for key, value in metadata.items():
                    story.append(Paragraph(f"<b>{key}:</b> {value}", styles['Normal']))
                story.append(Spacer(1, 12))
            
            # Add sections
            for section in sorted(sections, key=lambda x: x.order):
                # Section title
                story.append(Paragraph(section.title, styles['Heading2']))
                story.append(Spacer(1, 6))
                
                # Section content
                if section.content:
                    story.append(Paragraph(section.content, styles['Normal']))
                    story.append(Spacer(1, 12))
                
                # Add data tables if present
                if section.data and section.section_type == "table":
                    table_data = self._format_table_data(section.data)
                    if table_data:
                        table = Table(table_data)
                        table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 12),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black)
                        ]))
                        story.append(table)
                        story.append(Spacer(1, 12))
            
            # Build PDF
            doc.build(story)
            
            # Get PDF bytes
            pdf_bytes = buffer.getvalue()
            buffer.close()
            
            self.logger.info(f"Generated PDF report with {len(sections)} sections")
            return pdf_bytes
            
        except Exception as e:
            self.logger.error(f"Error generating PDF report: {e}")
            raise
    
    def _format_table_data(self, data: Dict[str, Any]) -> List[List[str]]:
        """Format data for table display"""
        if not data:
            return []
        
        if "headers" in data and "rows" in data:
            table_data = [data["headers"]]
            table_data.extend(data["rows"])
            return table_data
        
        # Convert dict to table format
        if isinstance(data, dict):
            headers = ["Key", "Value"]
            rows = [[str(k), str(v)] for k, v in data.items()]
            return [headers] + rows
        
        return []


class ExcelReportGenerator:
    """Excel report generation using OpenPyXL"""
    
    def __init__(self):
        self.logger = logging.getLogger(f"{__name__}.{self.__class__.__name__}")
    
    async def generate_excel_report(
        self, 
        sections: List[ReportSection], 
        title: str,
        metadata: Dict[str, Any] = None
    ) -> bytes:
        """Generate an Excel report from sections"""
        
        if not OPENPYXL_AVAILABLE:
            raise ImportError("OpenPyXL is required for Excel generation")
        
        try:
            workbook = Workbook()
            
            # Remove default sheet
            workbook.remove(workbook.active)
            
            # Create summary sheet
            summary_sheet = workbook.create_sheet("Summary")
            
            # Add title
            summary_sheet['A1'] = title
            summary_sheet['A1'].font = Font(size=16, bold=True)
            summary_sheet.merge_cells('A1:D1')
            
            row = 3
            
            # Add metadata
            if metadata:
                summary_sheet[f'A{row}'] = "Report Information"
                summary_sheet[f'A{row}'].font = Font(bold=True)
                row += 1
                
                for key, value in metadata.items():
                    summary_sheet[f'A{row}'] = key
                    summary_sheet[f'B{row}'] = str(value)
                    row += 1
                
                row += 1
            
            # Add section summaries
            summary_sheet[f'A{row}'] = "Sections"
            summary_sheet[f'A{row}'].font = Font(bold=True)
            row += 1
            
            for section in sorted(sections, key=lambda x: x.order):
                summary_sheet[f'A{row}'] = section.title
                summary_sheet[f'B{row}'] = section.section_type
                row += 1
            
            # Create individual sheets for sections with data
            for section in sections:
                if section.data and section.section_type in ["table", "data"]:
                    sheet_name = section.title[:31]  # Excel sheet name limit
                    sheet = workbook.create_sheet(sheet_name)
                    
                    # Add section title
                    sheet['A1'] = section.title
                    sheet['A1'].font = Font(size=14, bold=True)
                    
                    # Add data
                    if isinstance(section.data, dict):
                        if "headers" in section.data and "rows" in section.data:
                            # Table format
                            self._add_table_to_sheet(sheet, section.data, start_row=3)
                        else:
                            # Key-value format
                            self._add_dict_to_sheet(sheet, section.data, start_row=3)
            
            # Save to bytes
            buffer = io.BytesIO()
            workbook.save(buffer)
            excel_bytes = buffer.getvalue()
            buffer.close()
            
            self.logger.info(f"Generated Excel report with {len(sections)} sections")
            return excel_bytes
            
        except Exception as e:
            self.logger.error(f"Error generating Excel report: {e}")
            raise
    
    def _add_table_to_sheet(self, sheet, data: Dict[str, Any], start_row: int = 1):
        """Add table data to Excel sheet"""
        headers = data.get("headers", [])
        rows = data.get("rows", [])
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=start_row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add data rows
        for row_idx, row_data in enumerate(rows, start_row + 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                sheet.cell(row=row_idx, column=col_idx).value = cell_value
    
    def _add_dict_to_sheet(self, sheet, data: Dict[str, Any], start_row: int = 1):
        """Add dictionary data to Excel sheet"""
        # Headers
        sheet.cell(row=start_row, column=1).value = "Key"
        sheet.cell(row=start_row, column=2).value = "Value"
        
        # Make headers bold
        sheet.cell(row=start_row, column=1).font = Font(bold=True)
        sheet.cell(row=start_row, column=2).font = Font(bold=True)
        
        # Add data
        for row_idx, (key, value) in enumerate(data.items(), start_row + 1):
            sheet.cell(row=row_idx, column=1).value = str(key)
            sheet.cell(row=row_idx, column=2).value = str(value)


class HTMLReportGenerator:
    """HTML report generation with CSS styling"""
    
    def __init__(self):
        self.logger = logging.getLogger(f"{__name__}.{self.__class__.__name__}")
    
    async def generate_html_report(
        self, 
        sections: List[ReportSection], 
        title: str,
        metadata: Dict[str, Any] = None,
        custom_css: str = None
    ) -> str:
        """Generate an HTML report from sections"""
        
        try:
            html_parts = []
            
            # HTML header
            html_parts.append(self._get_html_header(title, custom_css))
            
            # Body start
            html_parts.append('<body>')
            html_parts.append('<div class="report-container">')
            
            # Title
            html_parts.append(f'<h1 class="report-title">{title}</h1>')
            
            # Metadata
            if metadata:
                html_parts.append('<div class="metadata-section">')
                html_parts.append('<h2>Report Information</h2>')
                html_parts.append('<table class="metadata-table">')
                for key, value in metadata.items():
                    html_parts.append(f'<tr><td><strong>{key}:</strong></td><td>{value}</td></tr>')
                html_parts.append('</table>')
                html_parts.append('</div>')
            
            # Table of contents
            if len(sections) > 1:
                html_parts.append('<div class="toc-section">')
                html_parts.append('<h2>Table of Contents</h2>')
                html_parts.append('<ul class="toc-list">')
                for section in sorted(sections, key=lambda x: x.order):
                    section_id = section.title.lower().replace(' ', '-')
                    html_parts.append(f'<li><a href="#{section_id}">{section.title}</a></li>')
                html_parts.append('</ul>')
                html_parts.append('</div>')
            
            # Sections
            for section in sorted(sections, key=lambda x: x.order):
                section_id = section.title.lower().replace(' ', '-')
                html_parts.append(f'<div class="section" id="{section_id}">')
                html_parts.append(f'<h2 class="section-title">{section.title}</h2>')
                
                if section.content:
                    html_parts.append(f'<div class="section-content">{section.content}</div>')
                
                # Add visualizations
                if section.visualization and section.visualization.get("image_base64"):
                    html_parts.append('<div class="visualization">')
                    img_src = f"data:image/png;base64,{section.visualization['image_base64']}"
                    html_parts.append(f'<img src="{img_src}" alt="Visualization" class="chart-image">')
                    html_parts.append('</div>')
                
                # Add data tables
                if section.data and section.section_type == "table":
                    html_parts.append(self._format_html_table(section.data))
                
                html_parts.append('</div>')
            
            # Body end
            html_parts.append('</div>')
            html_parts.append('</body>')
            html_parts.append('</html>')
            
            html_content = '\n'.join(html_parts)
            
            self.logger.info(f"Generated HTML report with {len(sections)} sections")
            return html_content
            
        except Exception as e:
            self.logger.error(f"Error generating HTML report: {e}")
            raise
    
    def _get_html_header(self, title: str, custom_css: str = None) -> str:
        """Generate HTML header with CSS"""
        default_css = """
        <style>
            body {
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                line-height: 1.6;
                margin: 0;
                padding: 20px;
                background-color: #f5f5f5;
            }
            .report-container {
                max-width: 1200px;
                margin: 0 auto;
                background-color: white;
                padding: 40px;
                border-radius: 10px;
                box-shadow: 0 0 20px rgba(0,0,0,0.1);
            }
            .report-title {
                color: #2c3e50;
                text-align: center;
                margin-bottom: 30px;
                border-bottom: 3px solid #3498db;
                padding-bottom: 10px;
            }
            .metadata-section {
                background-color: #ecf0f1;
                padding: 20px;
                border-radius: 5px;
                margin-bottom: 30px;
            }
            .metadata-table {
                width: 100%;
                border-collapse: collapse;
            }
            .metadata-table td {
                padding: 8px;
                border-bottom: 1px solid #bdc3c7;
            }
            .toc-section {
                background-color: #f8f9fa;
                padding: 20px;
                border-radius: 5px;
                margin-bottom: 30px;
            }
            .toc-list {
                list-style-type: none;
                padding-left: 0;
            }
            .toc-list li {
                margin-bottom: 10px;
            }
            .toc-list a {
                color: #3498db;
                text-decoration: none;
                font-weight: 500;
            }
            .toc-list a:hover {
                text-decoration: underline;
            }
            .section {
                margin-bottom: 40px;
                padding-bottom: 20px;
                border-bottom: 1px solid #ecf0f1;
            }
            .section-title {
                color: #2c3e50;
                margin-bottom: 15px;
            }
            .section-content {
                margin-bottom: 20px;
                text-align: justify;
            }
            .visualization {
                text-align: center;
                margin: 20px 0;
            }
            .chart-image {
                max-width: 100%;
                height: auto;
                border: 1px solid #ddd;
                border-radius: 5px;
            }
            .data-table {
                width: 100%;
                border-collapse: collapse;
                margin: 20px 0;
            }
            .data-table th,
            .data-table td {
                border: 1px solid #ddd;
                padding: 12px;
                text-align: left;
            }
            .data-table th {
                background-color: #3498db;
                color: white;
                font-weight: bold;
            }
            .data-table tr:nth-child(even) {
                background-color: #f2f2f2;
            }
            .data-table tr:hover {
                background-color: #e8f4fd;
            }
        </style>
        """
        
        css = custom_css if custom_css else default_css
        
        return f"""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>{title}</title>
            {css}
        </head>
        """
    
    def _format_html_table(self, data: Dict[str, Any]) -> str:
        """Format data as HTML table"""
        if not data:
            return ""
        
        table_html = ['<table class="data-table">']
        
        if "headers" in data and "rows" in data:
            # Add headers
            table_html.append('<thead><tr>')
            for header in data["headers"]:
                table_html.append(f'<th>{header}</th>')
            table_html.append('</tr></thead>')
            
            # Add rows
            table_html.append('<tbody>')
            for row in data["rows"]:
                table_html.append('<tr>')
                for cell in row:
                    table_html.append(f'<td>{cell}</td>')
                table_html.append('</tr>')
            table_html.append('</tbody>')
        else:
            # Convert dict to table
            table_html.append('<thead><tr><th>Key</th><th>Value</th></tr></thead>')
            table_html.append('<tbody>')
            for key, value in data.items():
                table_html.append(f'<tr><td>{key}</td><td>{value}</td></tr>')
            table_html.append('</tbody>')
        
        table_html.append('</table>')
        return '\n'.join(table_html)


class AdvancedReportingService:
    """Main advanced reporting service"""
    
    def __init__(self):
        self.visualization_engine = DataVisualizationEngine()
        self.pdf_generator = PDFReportGenerator()
        self.excel_generator = ExcelReportGenerator()
        self.html_generator = HTMLReportGenerator()
        self.templates = {}
        self.logger = logging.getLogger(f"{__name__}.{self.__class__.__name__}")
        
        # Initialize default templates
        self._initialize_default_templates()
    
    def _initialize_default_templates(self):
        """Initialize default report templates"""
        
        # Executive Summary Template
        executive_template = ReportTemplate(
            template_id="executive_summary",
            name="Executive Summary Report",
            description="High-level summary for executives and stakeholders",
            report_type=ReportType.EXECUTIVE_SUMMARY,
            sections=["overview", "key_metrics", "recommendations"],
            parameters={"include_charts": True, "detail_level": "high"}
        )
        self.templates["executive_summary"] = executive_template
        
        # Technical Analysis Template
        technical_template = ReportTemplate(
            template_id="technical_analysis",
            name="Technical Analysis Report",
            description="Detailed technical analysis with raw data",
            report_type=ReportType.TECHNICAL_REPORT,
            sections=["methodology", "data_analysis", "findings", "technical_details"],
            parameters={"include_raw_data": True, "detail_level": "comprehensive"}
        )
        self.templates["technical_analysis"] = technical_template
    
    async def generate_report(self, request: ReportRequest) -> Union[bytes, str]:
        """Generate a report based on the request"""
        
        try:
            self.logger.info(f"Generating {request.format.value} report: {request.title}")
            
            # Get template
            template = self.templates.get(request.template_id)
            if not template:
                raise ValueError(f"Template not found: {request.template_id}")
            
            # Prepare report sections
            sections = await self._prepare_report_sections(request, template)
            
            # Add visualizations if requested
            if request.include_visualizations:
                sections = await self._add_visualizations_to_sections(sections)
            
            # Generate report based on format
            if request.format == ReportFormat.PDF:
                return await self._generate_pdf_report(sections, request)
            elif request.format == ReportFormat.HTML:
                return await self._generate_html_report(sections, request)
            elif request.format == ReportFormat.EXCEL:
                return await self._generate_excel_report(sections, request)
            elif request.format == ReportFormat.JSON:
                return await self._generate_json_report(sections, request)
            elif request.format == ReportFormat.CSV:
                return await self._generate_csv_report(sections, request)
            else:
                raise ValueError(f"Unsupported report format: {request.format}")
                
        except Exception as e:
            self.logger.error(f"Error generating report: {e}")
            raise
    
    async def _prepare_report_sections(
        self, 
        request: ReportRequest, 
        template: ReportTemplate
    ) -> List[ReportSection]:
        """Prepare report sections based on template and data"""
        
        sections = []
        
        # Mock data preparation - in real implementation, this would fetch actual data
        mock_data = {
            "scan_results": [
                {"scanner": "email_validator", "status": "completed", "findings": "Valid email"},
                {"scanner": "phone_validator", "status": "completed", "findings": "Valid phone"},
                {"scanner": "social_media", "status": "completed", "findings": "3 profiles found"}
            ],
            "metrics": {
                "total_scans": 15,
                "successful_scans": 12,
                "success_rate": 80.0,
                "avg_execution_time": 2.5
            },
            "correlations": [
                {"entities": ["john@example.com", "+1234567890"], "strength": 0.85}
            ]
        }
        
        # Create sections based on template
        if "overview" in template.sections:
            overview_section = ReportSection(
                title="Executive Overview",
                content=f"Intelligence analysis completed for target with {mock_data['metrics']['success_rate']}% success rate. "
                       f"Total of {mock_data['metrics']['total_scans']} scanners executed with "
                       f"{mock_data['metrics']['successful_scans']} successful completions.",
                section_type="text",
                order=1
            )
            sections.append(overview_section)
        
        if "key_metrics" in template.sections:
            metrics_section = ReportSection(
                title="Key Metrics",
                content="Performance metrics for the intelligence gathering operation.",
                section_type="table",
                data={
                    "headers": ["Metric", "Value"],
                    "rows": [
                        ["Total Scans", str(mock_data['metrics']['total_scans'])],
                        ["Successful Scans", str(mock_data['metrics']['successful_scans'])],
                        ["Success Rate", f"{mock_data['metrics']['success_rate']}%"],
                        ["Average Execution Time", f"{mock_data['metrics']['avg_execution_time']}s"]
                    ]
                },
                order=2
            )
            sections.append(metrics_section)
        
        if "findings" in template.sections:
            findings_section = ReportSection(
                title="Key Findings",
                content="Detailed findings from the intelligence analysis.",
                section_type="table",
                data={
                    "headers": ["Scanner", "Status", "Findings"],
                    "rows": [
                        [result["scanner"], result["status"], result["findings"]]
                        for result in mock_data["scan_results"]
                    ]
                },
                order=3
            )
            sections.append(findings_section)
        
        if "recommendations" in template.sections:
            recommendations_section = ReportSection(
                title="Recommendations",
                content="Based on the analysis results, we recommend: "
                       "1. Continue monitoring the identified entities. "
                       "2. Cross-reference findings with additional data sources. "
                       "3. Implement additional security measures if threats are identified.",
                section_type="text",
                order=4
            )
            sections.append(recommendations_section)
        
        return sections
    
    async def _add_visualizations_to_sections(
        self, 
        sections: List[ReportSection]
    ) -> List[ReportSection]:
        """Add visualizations to appropriate sections"""
        
        for section in sections:
            if section.section_type == "table" and section.data:
                # Create visualization based on data
                if section.title == "Key Metrics":
                    # Create bar chart for metrics
                    chart_data = {
                        "categories": [row[0] for row in section.data["rows"][:3]],  # First 3 metrics
                        "values": [float(str(row[1]).rstrip('%s')) for row in section.data["rows"][:3]]
                    }
                    
                    image_base64 = await self.visualization_engine.create_visualization(
                        chart_data,
                        VisualizationType.BAR_CHART,
                        "Key Performance Metrics"
                    )
                    
                    if image_base64:
                        section.visualization = {"image_base64": image_base64}
        
        return sections
    
    async def _generate_pdf_report(
        self, 
        sections: List[ReportSection], 
        request: ReportRequest
    ) -> bytes:
        """Generate PDF report"""
        metadata = {
            "Generated": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"),
            "Report Type": request.report_type.value,
            "Template": request.template_id
        }
        
        return await self.pdf_generator.generate_pdf_report(
            sections, request.title, metadata
        )
    
    async def _generate_html_report(
        self, 
        sections: List[ReportSection], 
        request: ReportRequest
    ) -> str:
        """Generate HTML report"""
        metadata = {
            "Generated": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"),
            "Report Type": request.report_type.value,
            "Template": request.template_id
        }
        
        return await self.html_generator.generate_html_report(
            sections, request.title, metadata, request.custom_styling
        )
    
    async def _generate_excel_report(
        self, 
        sections: List[ReportSection], 
        request: ReportRequest
    ) -> bytes:
        """Generate Excel report"""
        metadata = {
            "Generated": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"),
            "Report Type": request.report_type.value,
            "Template": request.template_id
        }
        
        return await self.excel_generator.generate_excel_report(
            sections, request.title, metadata
        )
    
    async def _generate_json_report(
        self, 
        sections: List[ReportSection], 
        request: ReportRequest
    ) -> str:
        """Generate JSON report"""
        report_data = {
            "title": request.title,
            "report_type": request.report_type.value,
            "template_id": request.template_id,
            "generated_at": datetime.utcnow().isoformat(),
            "sections": [
                {
                    "title": section.title,
                    "content": section.content,
                    "type": section.section_type,
                    "data": section.data,
                    "order": section.order
                }
                for section in sections
            ]
        }
        
        return json.dumps(report_data, indent=2, default=str)
    
    async def _generate_csv_report(
        self, 
        sections: List[ReportSection], 
        request: ReportRequest
    ) -> str:
        """Generate CSV report (for tabular data)"""
        csv_lines = []
        csv_lines.append(f"Report Title,{request.title}")
        csv_lines.append(f"Generated,{datetime.utcnow().isoformat()}")
        csv_lines.append("")
        
        for section in sections:
            csv_lines.append(f"Section,{section.title}")
            
            if section.data and "headers" in section.data and "rows" in section.data:
                # Add headers
                csv_lines.append(",".join(section.data["headers"]))
                
                # Add rows
                for row in section.data["rows"]:
                    csv_lines.append(",".join(str(cell) for cell in row))
            else:
                csv_lines.append(f"Content,{section.content}")
            
            csv_lines.append("")
        
        return "\n".join(csv_lines)
    
    def register_template(self, template: ReportTemplate):
        """Register a new report template"""
        self.templates[template.template_id] = template
        self.logger.info(f"Registered report template: {template.name}")
    
    def get_available_templates(self) -> List[Dict[str, Any]]:
        """Get list of available report templates"""
        return [
            {
                "template_id": template.template_id,
                "name": template.name,
                "description": template.description,
                "type": template.report_type.value,
                "sections": template.sections
            }
            for template in self.templates.values()
        ]


# Global reporting service instance
reporting_service = AdvancedReportingService()


# Convenience functions
async def generate_intelligence_report(
    title: str,
    template_id: str = "executive_summary",
    format: ReportFormat = ReportFormat.HTML,
    include_visualizations: bool = True
) -> Union[bytes, str]:
    """Generate an intelligence report"""
    
    request = ReportRequest(
        report_id=f"report_{int(datetime.utcnow().timestamp())}",
        template_id=template_id,
        report_type=ReportType.EXECUTIVE_SUMMARY,
        format=format,
        title=title,
        include_visualizations=include_visualizations
    )
    
    return await reporting_service.generate_report(request)


async def create_visualization(
    data: Dict[str, Any],
    chart_type: VisualizationType,
    title: str = ""
) -> Optional[str]:
    """Create a data visualization"""
    return await reporting_service.visualization_engine.create_visualization(
        data, chart_type, title
    )